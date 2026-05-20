from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ID_HEADER_CANDIDATES = {"id_entree_original", "id_entree", "id", "numero", "numéro"}


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    return (
        text.strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
        .replace("__", "_")
    )


def normalize_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return f"{value:03d}"
    if isinstance(value, float) and value.is_integer():
        return f"{int(value):03d}"

    text = str(value).strip()
    if not text:
        return ""
    if text.isdigit():
        return f"{int(text):03d}"
    digits = "".join(ch for ch in text if ch.isdigit())
    if digits:
        return f"{int(digits):03d}"
    return text


def _row_values(row: tuple[Any, ...]) -> list[Any]:
    return [cell for cell in row]


def _find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows):
        headers = [normalize_header(value) for value in _row_values(row)]
        if any(header in ID_HEADER_CANDIDATES for header in headers):
            return index, headers
    raise ValueError("No journal-like header row found: missing ID column")


def load_journal_records(path: str | Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    """Load journal rows from an XLSX file without mutating it.

    The reader is intentionally permissive: it searches for a header row containing an ID-like
    column, then returns normalized dictionaries. It does not infer documentary truth from the
    journal; it only reads treatment-state rows.
    """

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = [sheet_name] if sheet_name else workbook.sheetnames
    records: list[dict[str, Any]] = []

    for name in sheets:
        worksheet = workbook[name]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            continue
        try:
            header_index, headers = _find_header_row(rows)
        except ValueError:
            continue

        for row in rows[header_index + 1 :]:
            values = _row_values(row)
            if not any(value not in (None, "") for value in values):
                continue
            record = {
                headers[col_index]: values[col_index] if col_index < len(values) else None
                for col_index in range(len(headers))
                if headers[col_index]
            }
            record["_sheet"] = name
            record["_row_index"] = header_index + 2
            records.append(record)

    if not records:
        raise ValueError("No journal records found in workbook")
    return records


def index_records_by_id(records: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    index: dict[str, dict[str, Any]] = {}
    for record in records:
        id_value = ""
        for key in ID_HEADER_CANDIDATES:
            if key in record:
                id_value = normalize_id(record[key])
                break
        if id_value:
            index[id_value] = record
    return index


def record_text(record: dict[str, Any] | None) -> str:
    if not record:
        return ""
    return " ".join("" if value is None else str(value) for value in record.values()).lower()


@dataclass(frozen=True)
class Post005GuardReport:
    id_026_guard_ok: bool
    id_030_guard_ok: bool
    id_032_guard_ok: bool
    details: dict[str, str]

    @property
    def ok(self) -> bool:
        return self.id_026_guard_ok and self.id_030_guard_ok and self.id_032_guard_ok


def check_post005_guards(records: list[dict[str, Any]]) -> Post005GuardReport:
    """Check the known post-005 steering invariants.

    Guards:
    - 026 and 030 must not look like new/untreated entries.
    - 032 must not look like RUN_001/RUN_002 done.
    """

    indexed = index_records_by_id(records)
    text_026 = record_text(indexed.get("026"))
    text_030 = record_text(indexed.get("030"))
    text_032 = record_text(indexed.get("032"))

    id_026_guard_ok = bool(text_026) and "jamais_traite" not in text_026 and "jamais traité" not in text_026
    id_030_guard_ok = bool(text_030) and "jamais_traite" not in text_030 and "jamais traité" not in text_030
    id_032_guard_ok = "run_001" not in text_032 and "run_002" not in text_032

    return Post005GuardReport(
        id_026_guard_ok=id_026_guard_ok,
        id_030_guard_ok=id_030_guard_ok,
        id_032_guard_ok=id_032_guard_ok,
        details={
            "026": text_026,
            "030": text_030,
            "032": text_032,
        },
    )


__all__ = [
    "Post005GuardReport",
    "check_post005_guards",
    "index_records_by_id",
    "load_journal_records",
    "normalize_header",
    "normalize_id",
    "record_text",
]
